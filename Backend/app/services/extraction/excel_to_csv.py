"""In-memory Excel (.xlsx) → CSV conversion.

Excel files are binary (zipped XML). Feeding their raw bytes to the CSV
parser or the CSV→PDF renderer produces garbage. We use openpyxl to read
every sheet and emit one CSV. If the workbook has multiple sheets, sheets
are concatenated with a blank-line separator and a `# sheet: <name>`
comment row so downstream extractors (deterministic parser or Gemini)
can still see section boundaries.

The function is intentionally tolerant: a malformed workbook raises
`ExcelDecodeError`, which the caller surfaces as a friendly upload-failure
note rather than a 500.
"""

from __future__ import annotations

import csv as _csv
import io
import logging
from typing import Iterable


logger = logging.getLogger(__name__)


XLSX_MAGIC = b"PK\x03\x04"   # all xlsx files are zip archives
XLS_MAGIC_DOC = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"   # legacy .xls (Compound Document)


class ExcelDecodeError(ValueError):
    """Raised when openpyxl can't read the workbook."""


def looks_like_xlsx(sample: bytes) -> bool:
    """Magic-byte sniff. .xlsx is a zip; legacy .xls uses the Compound File
    binary format which openpyxl can't read (we surface a clear error for it).
    """
    return sample.startswith(XLSX_MAGIC)


def looks_like_legacy_xls(sample: bytes) -> bool:
    return sample.startswith(XLS_MAGIC_DOC)


def xlsx_bytes_to_csv(content: bytes) -> bytes:
    """Read every sheet in the workbook and serialise to CSV bytes (utf-8).
    Empty trailing rows are stripped; per-cell newlines are preserved
    inside quoted fields by csv.writer."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as e:  # pragma: no cover — guarded by requirements.txt
        raise ExcelDecodeError(
            "openpyxl is not installed. Run `pip install openpyxl`."
        ) from e

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ExcelDecodeError(
            f"openpyxl could not open the .xlsx file: {e}. "
            "It may be a legacy .xls (binary Compound File) or password-protected."
        ) from e

    buf = io.StringIO()
    writer = _csv.writer(buf)
    sheet_names = wb.sheetnames or []
    for i, name in enumerate(sheet_names):
        sheet = wb[name]
        if i > 0:
            writer.writerow([])
        if len(sheet_names) > 1:
            # Section marker the rule-based parser ignores and Gemini reads.
            writer.writerow([f"# sheet: {name}"])
        rows = _iter_rows(sheet)
        for row in rows:
            writer.writerow(row)
    text = buf.getvalue()
    return text.encode("utf-8")


def _iter_rows(sheet) -> Iterable[list[str]]:
    for row in sheet.iter_rows(values_only=True):
        if not any(c not in (None, "") for c in row):
            continue
        yield ["" if v is None else str(v) for v in row]
